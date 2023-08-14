import csv
# biblioteca que se conecta com a AWS
import boto3
import openpyxl

# Cria um cliente da AWS lendo as credenciais do arquivo credentials.csv
with open('credentials.csv', 'r') as input:
    next(input)
    reader = csv.reader(input)
    for line in reader:
        access_key_id = line[2]
        secret_access_key = line[3]

# Cria o cliente de acesso a API Rekognition da AWS
client = boto3.client('rekognition',
                      aws_access_key_id=access_key_id,
                      aws_secret_access_key=secret_access_key,
                      region_name='us-east-1')

# Lista das fotos a serem processadas
photo_list = ["foto1.jpg",
              "foto2.jpg",
              "foto3.jpg",
              "foto4.jpg",
              "foto5.jpg",
              "foto6.jpg",
              "foto7.jpeg",
              "foto8.jpg",
              "foto9.jpg",
              "foto10.jpeg",
              "foto11.jpg",
              "foto12.jfif",
              "foto13.jpg",
              "foto14.jfif",
              "foto15.jpg",]

# Tenta abrir o arquivo .xlsx, cria um novo se não existir
try:
    workbook = openpyxl.load_workbook("resultado.xlsx")
    worksheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    # Escreve os cabeçalhos das colunas
    worksheet.cell(row=1, column=1, value="Label (Rótulo)")
    worksheet.cell(row=1, column=2, value="Confidence (Confiança) em %")
    worksheet.cell(row=1, column=3, value="Detecção de rosto")
    worksheet.cell(row=1, column=4, value="Foto")

row = worksheet.max_row + 1

# Processa cada foto da lista
for photo in photo_list:
    # Converte a imagem em um array de bytes
    with open(photo, 'rb') as source_image:
        source_bytes = source_image.read()

    # Usa o cliente para detectar as labels, melhor configuração de camadas e confiança mínima
    response = client.detect_labels(Image={'Bytes': source_bytes},
                                    MaxLabels=20,
                                    MinConfidence=80)

    # Verifica se há rostos na foto/ serviço independente
    face_response = client.detect_faces(Image={'Bytes': source_bytes}, Attributes=["ALL"])
    has_face = False
    if len(face_response["FaceDetails"]) > 0:
        has_face = True
        print("Foram encontrados rostos na foto.")
    else:
        print("Não foram encontradas rostos na foto.")

    # Adiciona uma nova linha para cada label detectada
    for i, label in enumerate(response["Labels"]):
        worksheet.cell(row=row, column=1, value=label["Name"])
        worksheet.cell(row=row, column=2, value=label["Confidence"])
        worksheet.cell(row=row, column=3, value=has_face)
        worksheet.cell(row=row, column=4, value=photo.split("/")[-1])
        row += 1

# Salva o arquivo .xlsx
print("Salvando o arquivo...")
workbook.save("resultado.xlsx")
print("Arquivo salvo com sucesso!")
